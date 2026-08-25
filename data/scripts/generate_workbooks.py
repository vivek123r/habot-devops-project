# -*- coding: utf-8 -*-
"""
Candidate : Vivek R
Contact   : vivekravi9496497657@gmail.com | +91 8590609366
Project   : HabotConnect Hiring Project 1.0
            Junior Cloud & DevOps Engineer (GCP / Django / React)
File      : data/scripts/generate_workbooks.py
Purpose   : Reproducibly builds the two submission workbooks demanded by the
            hiring document:

              schema_mapping.xlsx    - payload field to BigQuery column mapping
              dcyn_logic_matrix.xlsx - Decision Yes/No decision and validation matrices

          Compliance properties (hiring document, submission instruction 3c):
            - Every worksheet has Wrap Text enabled on every populated cell,
              verified programmatically after generation.
            - Full forms only: no slang, no abbreviations, no placeholders.
              Where a short identifier exists (R01, R02) it appears together
              with its complete description.

          Usage:  python data/scripts/generate_workbooks.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[2] / "backend"))

from onboarding import constants  # noqa: E402 - path bootstrap above

REPO_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = REPO_ROOT / "data"

HEADER_FILL = PatternFill("solid", fgColor="1F4E5F")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP_TOP_LEFT = Alignment(wrap_text=True, vertical="top", horizontal="left")


def build_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[list],
    widths: list[int],
) -> None:
    sheet = workbook.create_sheet(title=title)
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_cells = sheet[1]
    for index, cell in enumerate(header_cells, start=1):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP_TOP_LEFT
        sheet.column_dimensions[get_column_letter(index)].width = widths[index - 1]

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP_LEFT

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def enforce_wrap_text_everywhere(path: Path) -> None:
    """Post-generation verification: hiring rule 3c is load-bearing."""
    workbook = load_workbook(path)
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    assert (
                        cell.alignment.wrap_text is True
                    ), f"{path.name}/{sheet.title}/{cell.coordinate} lost Wrap Text"


def main() -> None:
    # ------------------------------------------------------------------
    # Workbook one: schema_mapping.xlsx
    # ------------------------------------------------------------------
    schema_rows = [
        [
            "parent_full_name",
            "Text",
            "Yes",
            constants.PARENT_FULL_NAME_MIN_LENGTH,
            constants.PARENT_FULL_NAME_MAX_LENGTH,
            "Letters of any alphabet separated by single spaces, hyphens, apostrophes, or periods; trimmed before checks.",
            "ERR_REQUIRED_PARENT_FULL_NAME, ERR_MIN_LENGTH_PARENT_FULL_NAME, ERR_MAX_LENGTH_PARENT_FULL_NAME, ERR_GRAMMAR_PARENT_FULL_NAME, ERR_EXACT_JSON_TYPE_STRING_REQUIRED",
            "parent_full_name",
            "STRING",
            "REQUIRED",
        ],
        [
            "parent_email",
            "Text",
            "Yes",
            3,
            constants.EMAIL_MAX_LENGTH,
            "Exactly one commercial at sign, non-empty local part, domain with at least one dot; lowercased before storage.",
            "ERR_REQUIRED_PARENT_EMAIL, ERR_INVALID_PARENT_EMAIL, ERR_MAX_LENGTH_PARENT_EMAIL",
            "parent_email",
            "STRING",
            "REQUIRED",
        ],
        [
            "parent_phone",
            "Text",
            "Yes",
            9,
            16,
            "International direct dialing format: plus sign, country code without leading zero, seven to fifteen following digits.",
            "ERR_REQUIRED_PARENT_PHONE, ERR_FORMAT_PARENT_PHONE, ERR_EXACT_JSON_TYPE_STRING_REQUIRED",
            "parent_phone",
            "STRING",
            "REQUIRED",
        ],
        [
            "child_full_name",
            "Text",
            "Yes",
            constants.CHILD_FULL_NAME_MIN_LENGTH,
            constants.CHILD_FULL_NAME_MAX_LENGTH,
            "Identical grammar to the parent full name field.",
            "ERR_REQUIRED_CHILD_FULL_NAME, ERR_MIN_LENGTH_CHILD_FULL_NAME, ERR_MAX_LENGTH_CHILD_FULL_NAME, ERR_GRAMMAR_CHILD_FULL_NAME",
            "child_full_name",
            "STRING",
            "REQUIRED",
        ],
        [
            "child_date_of_birth",
            "Calendar date",
            "Yes",
            "1900-01-01",
            "Evaluation date",
            "ISO calendar date; completed age derived from it must fall inside the inclusive two-to-eighteen window.",
            "ERR_REQUIRED_CHILD_DATE_OF_BIRTH, ERR_FORMAT_CHILD_DATE_OF_BIRTH, ERR_AGE_OUTSIDE_WINDOW_CHILD_DATE_OF_BIRTH",
            "child_date_of_birth",
            "DATE",
            "REQUIRED",
        ],
        [
            "learning_difficulty_category",
            "Enumeration",
            "Yes",
            "One of seven values",
            "Longest member length sixty-four",
            "Exact membership in the closed seven-value category set; case sensitive; abbreviations rejected.",
            "ERR_REQUIRED_LEARNING_DIFFICULTY_CATEGORY, ERR_UNKNOWN_LEARNING_DIFFICULTY_CATEGORY",
            "learning_difficulty_category",
            "STRING",
            "REQUIRED",
        ],
        [
            "support_needs_summary",
            "Text",
            "No",
            constants.SUPPORT_NEEDS_SUMMARY_MIN_LENGTH,
            constants.SUPPORT_NEEDS_SUMMARY_MAX_LENGTH,
            "Optional context; empty after trimming is treated as absent; otherwise trimmed length must sit inside bounds.",
            "ERR_MIN_LENGTH_SUPPORT_NEEDS_SUMMARY, ERR_MAX_LENGTH_SUPPORT_NEEDS_SUMMARY",
            "support_needs_summary",
            "STRING",
            "NULLABLE",
        ],
        [
            "parental_consent_granted",
            "Boolean",
            "Yes",
            "true",
            "true",
            "Decision Yes/No consent gate: exactly the JSON boolean true; false, null, numbers, and strings all reject.",
            "ERR_REQUIRED_PARENTAL_CONSENT_GRANTED, ERR_CONSENT_VALUE_FALSE_REJECTED, ERR_EXACT_JSON_TYPE_BOOLEAN_REQUIRED",
            "parental_consent_granted",
            "BOOLEAN",
            "REQUIRED",
        ],
        [
            "data_processing_consent_granted",
            "Boolean",
            "Yes",
            "true",
            "true",
            "Second Decision Yes/No consent gate with identical strictness.",
            "ERR_REQUIRED_DATA_PROCESSING_CONSENT_GRANTED, ERR_CONSENT_VALUE_FALSE_REJECTED, ERR_EXACT_JSON_TYPE_BOOLEAN_REQUIRED",
            "data_processing_consent_granted",
            "BOOLEAN",
            "REQUIRED",
        ],
        [
            "(server generated)",
            "Universally unique identifier",
            "Generated on acceptance",
            "Not applicable",
            "Not applicable",
            "Created by the application programming interface at persistence time; never accepted from callers.",
            "Not applicable",
            "submission_id",
            "STRING",
            "REQUIRED",
        ],
        [
            "(server generated)",
            "Audit record",
            "Generated on acceptance",
            "Not applicable",
            "Not applicable",
            "Complete Decision Yes/No evaluation record retained for audit.",
            "Not applicable",
            "dcyn_decision_record",
            "JSON",
            "NULLABLE",
        ],
        [
            "(server generated)",
            "Aggregate verdict",
            "Generated on acceptance",
            "Not applicable",
            "Not applicable",
            "True only when every Decision Yes/No rule returned Yes.",
            "Not applicable",
            "dcyn_all_rules_passed",
            "BOOLEAN",
            "REQUIRED",
        ],
        [
            "(server generated)",
            "Timestamp",
            "Generated on acceptance",
            "Not applicable",
            "Not applicable",
            "Coordinated Universal Time moment the form passed validation.",
            "Not applicable",
            "submitted_at",
            "TIMESTAMP",
            "REQUIRED",
        ],
    ]

    limits_rows = [
        [
            "PARENT_FULL_NAME_MIN_LENGTH",
            "Minimum character count",
            str(constants.PARENT_FULL_NAME_MIN_LENGTH),
        ],
        [
            "PARENT_FULL_NAME_MAX_LENGTH",
            "Maximum character count",
            str(constants.PARENT_FULL_NAME_MAX_LENGTH),
        ],
        [
            "CHILD_FULL_NAME_MIN_LENGTH",
            "Minimum character count",
            str(constants.CHILD_FULL_NAME_MIN_LENGTH),
        ],
        [
            "CHILD_FULL_NAME_MAX_LENGTH",
            "Maximum character count",
            str(constants.CHILD_FULL_NAME_MAX_LENGTH),
        ],
        [
            "EMAIL_MAX_LENGTH",
            "Maximum character count",
            str(constants.EMAIL_MAX_LENGTH),
        ],
        [
            "SUPPORT_NEEDS_SUMMARY_MIN_LENGTH",
            "Minimum character count",
            str(constants.SUPPORT_NEEDS_SUMMARY_MIN_LENGTH),
        ],
        [
            "SUPPORT_NEEDS_SUMMARY_MAX_LENGTH",
            "Maximum character count",
            str(constants.SUPPORT_NEEDS_SUMMARY_MAX_LENGTH),
        ],
        [
            "CHILD_AGE_MIN_YEARS",
            "Minimum completed years",
            str(constants.CHILD_AGE_MIN_YEARS),
        ],
        [
            "CHILD_AGE_MAX_YEARS",
            "Maximum completed years",
            str(constants.CHILD_AGE_MAX_YEARS),
        ],
    ]

    schema_book = Workbook()
    schema_book.remove(schema_book.active)
    build_sheet(
        schema_book,
        "Schema Mapping",
        [
            "Payload Field Name",
            "Data Type",
            "Required",
            "Lower Bound",
            "Upper Bound",
            "Validation Rule Description",
            "Machine Readable Error Codes",
            "BigQuery Column",
            "BigQuery Type",
            "BigQuery Mode",
        ],
        schema_rows,
        [26, 14, 12, 16, 18, 46, 46, 24, 12, 10],
    )
    build_sheet(
        schema_book,
        "Field Limits",
        ["Source Constant Identifier", "Limit Kind", "Limit Value"],
        limits_rows,
        [42, 30, 20],
    )
    schema_path = DATA_DIR / "schema_mapping.xlsx"
    schema_book.save(schema_path)
    enforce_wrap_text_everywhere(schema_path)

    # ------------------------------------------------------------------
    # Workbook two: dcyn_logic_matrix.xlsx
    # ------------------------------------------------------------------
    decision_rows = [
        [
            "R01",
            "Parent full name validity",
            "parent_full_name",
            "Present, two to one hundred twenty characters, matches letter grammar",
            "Absent, mistyped, wrong length, or fails grammar",
            "Continue to next rule",
            "Reject submission; return error code",
        ],
        [
            "R02",
            "Parent email structural validity",
            "parent_email",
            "Exactly one commercial at sign, valid local part and domain, within two hundred fifty four characters",
            "Any structural defect or absence",
            "Continue to next rule",
            "Reject submission; return error code",
        ],
        [
            "R03",
            "Parent phone format",
            "parent_phone",
            "Matches international direct dialing pattern",
            "Missing plus sign, leading zero country code, wrong digit count, or absent",
            "Continue to next rule",
            "Reject submission; return error code",
        ],
        [
            "R04",
            "Child full name validity",
            "child_full_name",
            "Same grammar and window as parent name",
            "Any defect or absence",
            "Continue to next rule",
            "Reject submission; return error code",
        ],
        [
            "R05",
            "Child date of birth calendar validity",
            "child_date_of_birth",
            "Parses as ISO calendar date",
            "Wrong shape, impossible date, or absent",
            "Enable R06 evaluation",
            "Reject submission; R06 reported as not evaluable",
        ],
        [
            "R06",
            "Child completed age window",
            "derived from child_date_of_birth",
            f"Completed years between {constants.CHILD_AGE_MIN_YEARS} and {constants.CHILD_AGE_MAX_YEARS} inclusive on the evaluation date",
            "Younger than two, older than eighteen, or not evaluable",
            "Continue to next rule",
            "Reject submission; return age window error code",
        ],
        [
            "R07",
            "Learning difficulty category membership",
            "learning_difficulty_category",
            "Equals exactly one of the seven documented values",
            "Unknown value, abbreviation, case mismatch, or absent",
            "Continue to next rule",
            "Reject submission; return unknown category code",
        ],
        [
            "R08",
            "Support needs summary bound check",
            "support_needs_summary",
            "Absent, empty after trimming, or between ten and one thousand characters",
            "Present non-empty outside the bounds",
            "Continue to next rule",
            "Reject submission; return length error code",
        ],
        [
            "R09",
            "Parental consent gate",
            "parental_consent_granted",
            "Exactly the boolean true",
            "False, null, number, string, or absent",
            "Continue to next rule",
            "Reject submission; consent cannot be assumed",
        ],
        [
            "R10",
            "Data processing consent gate",
            "data_processing_consent_granted",
            "Exactly the boolean true",
            "False, null, number, string, or absent",
            "Submission accepted",
            "Reject submission; consent cannot be assumed",
        ],
        [
            "R11",
            "Payload envelope strictness",
            "entire payload object",
            "Contains only the nine documented field names",
            "Any undocumented key present",
            "Continue evaluating remaining rules",
            "Reject submission; unknown fields listed in detail",
        ],
        [
            "R12",
            "Exact type conformance",
            "every present field",
            "Each present field has exactly its documented JSON type",
            "Any coercion candidate present (number where text belongs, string where boolean belongs)",
            "Continue to next rule",
            "Reject submission; offending fields listed",
        ],
    ]

    validation_rows = [
        [
            "V01",
            "Complete correct payload",
            "All nine documented fields with valid values",
            "Accept",
            "None",
            "201",
            "Every rule evaluates Yes",
        ],
        [
            "V02",
            "Missing parent full name",
            "Field removed from payload",
            "Reject",
            "ERR_REQUIRED_PARENT_FULL_NAME",
            "400",
            "Presence is binary",
        ],
        [
            "V03",
            "Missing parental consent",
            "Field removed from payload",
            "Reject",
            "ERR_REQUIRED_PARENTAL_CONSENT_GRANTED",
            "400",
            "Consent can never be defaulted",
        ],
        [
            "V04",
            "Consent supplied as text",
            '"true" as a string',
            "Reject",
            "ERR_EXACT_JSON_TYPE_BOOLEAN_REQUIRED",
            "400",
            "Coercion forbidden by exact-type rule",
        ],
        [
            "V05",
            "Consent supplied as number",
            "1",
            "Reject",
            "ERR_EXACT_JSON_TYPE_BOOLEAN_REQUIRED",
            "400",
            "Coercion forbidden by exact-type rule",
        ],
        [
            "V06",
            "Consent explicitly refused",
            "false",
            "Reject",
            "ERR_CONSENT_VALUE_FALSE_REJECTED",
            "400",
            "Only literal true passes the gate",
        ],
        [
            "V07",
            "Age exactly two years on evaluation date",
            "Birth date equal to evaluation date minus two years",
            "Accept",
            "None",
            "201",
            "Inclusive boundary arithmetic",
        ],
        [
            "V08",
            "Age one day under two years",
            "Birth date one day later than the V07 birth date",
            "Reject",
            "ERR_AGE_OUTSIDE_WINDOW_CHILD_DATE_OF_BIRTH",
            "400",
            "Inclusive boundary arithmetic",
        ],
        [
            "V09",
            "Age exactly eighteen years",
            "Birth date equal to evaluation date minus eighteen years",
            "Accept",
            "None",
            "201",
            "Inclusive upper boundary",
        ],
        [
            "V10",
            "Age nineteen or older",
            "Birth date making completed age nineteen",
            "Reject",
            "ERR_AGE_OUTSIDE_WINDOW_CHILD_DATE_OF_BIRTH",
            "400",
            "Outside inclusive upper boundary",
        ],
        [
            "V11",
            "Impossible calendar date",
            "Thirtieth day of February",
            "Reject",
            "ERR_FORMAT_CHILD_DATE_OF_BIRTH",
            "400",
            "Calendar validation rejects non-existent dates",
        ],
        [
            "V12",
            "Number supplied for telephone",
            "Digits without quotes",
            "Reject",
            "ERR_EXACT_JSON_TYPE_STRING_REQUIRED",
            "400",
            "Type law precedes format law",
        ],
        [
            "V13",
            "Abbreviated category value",
            "Three-letter abbreviation",
            "Reject",
            "ERR_UNKNOWN_LEARNING_DIFFICULTY_CATEGORY",
            "400",
            "Full forms only policy",
        ],
        [
            "V14",
            "Undocumented extra field",
            "Additional key added to payload",
            "Reject",
            "ERR_UNKNOWN_FIELD_NOT_IN_DOCUMENTED_ENVELOPE",
            "400",
            "Envelope strictness rule",
        ],
        [
            "V15",
            "Telephone missing plus sign",
            "Country code without international prefix",
            "Reject",
            "ERR_FORMAT_PARENT_PHONE",
            "400",
            "Pattern match is binary",
        ],
        [
            "V16",
            "Summary one thousand one characters",
            "One character beyond maximum",
            "Reject",
            "ERR_MAX_LENGTH_SUPPORT_NEEDS_SUMMARY",
            "400",
            "Exclusive upper bound violation",
        ],
        [
            "V17",
            "Repeated identical submissions",
            "Same valid payload twice",
            "Accept both",
            "None",
            "201",
            "Deterministic validator yields identical decision records",
        ],
    ]

    logic_book = Workbook()
    logic_book.remove(logic_book.active)
    build_sheet(
        logic_book,
        "Decision Matrix",
        [
            "Rule Identifier",
            "Rule Description",
            "Input Under Evaluation",
            "Yes Outcome Condition",
            "No Outcome Condition",
            "Action When Yes",
            "Action When No",
        ],
        decision_rows,
        [14, 34, 28, 44, 44, 32, 40],
    )
    build_sheet(
        logic_book,
        "Validation Matrix",
        [
            "Test Case Identifier",
            "Scenario Description",
            "Input Value",
            "Expected Outcome",
            "Expected Machine Readable Code",
            "Expected Hypertext Transfer Protocol Status",
            "Deterministic Justification",
        ],
        validation_rows,
        [18, 36, 40, 16, 44, 22, 44],
    )
    logic_path = DATA_DIR / "dcyn_logic_matrix.xlsx"
    logic_book.save(logic_path)
    enforce_wrap_text_everywhere(logic_path)

    print(f"written: {schema_path}")
    print(f"written: {logic_path}")
    print("wrap text verification passed on every populated cell")


if __name__ == "__main__":
    main()
